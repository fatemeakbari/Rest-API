import json

from openpyxl import load_workbook

import datetime

from dashboard import db
from Models import Hero


def init_hero():

    wb = load_workbook(filename='/flaskapp/app/init_db/hero.xlsx', read_only=True)
    #wb = load_workbook(filename='/home/fateme/myfolder/docker/restplus3/app/init_db/hero.xlsx', read_only=True)
    ws = wb.active



    data = {}

    for row in ws.rows:



        data['hero_id'] = int(row[0].value)

        data['max_health'] = float(row[1].value)

        data['danage_per_hit'] = float(row[2].value)

        data['defence'] = float(row[3].value)

        data['energy'] = float(row[4].value)

        data['stamina'] = float(row[5].value)

        data['price'] = int(row[6].value)



        data['minimumRequiredLevel'] = int(row[7].value)

        data['recovery_rate'] = datetime.datetime.now() + datetime.timedelta(minutes=int(row[8].value))





        new_hero = Hero(data['hero_id'], data['max_health'], data['danage_per_hit'], data['defence'],

                        data['energy'],

                        data['stamina'],

                        data['price'],

                        data['minimumRequiredLevel'],

                        data['recovery_rate']

                        )



        hero = new_hero.save_to_db()



        if hero:

            print('ok')

        else:

            print('khyli khari gavi famidi:(')
